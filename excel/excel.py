from openpyxl import *
from openpyxl.styles.colors import *
from ExampleResults import *
from openpyxl.styles.alignment import *
from openpyxl.styles import *
from openpyxl.utils import *
from pyshorteners import *

shortener = Shortener()
expStringLi = []
eduStringLi = []
expTempString = ""
eduTempString = ""

wb = load_workbook("Template.xlsx")

ws = wb.active

# iterate over amount of profiles
for i in range(profileCount):
    # imports names, title, currentCOmpany, locations, and the links 1 person at a time
    row = str(i+3)
    ws['B' + row].value = names[i]
    ws['C' + row].value = currentPos[i]
    ws['D' + row].value = currentCo[i]
    ws['E' + row].value = locations[i]
    tempLinkStr = links[i]
    ws['H' + row].value = shortener.tinyurl.short(tempLinkStr)
    
    # creates a bulleted list for exerpience and edu using for loops to iterate over the different amount of experience members and experience elements that a profile may have
    for y in range(len(experience)):
        for z in range(len(experience[y])):
            tinyURLString = shortener.tinyurl.short(str(experience[y][z][3]))
            expElementString = " - Position: " + str(experience[y][z][0]) + "\n    - Company: " + str(experience[y][z][1]) + "\n    - Duration: " + str(experience[y][z][2]) + "\n    - URL: " + tinyURLString

            expStringLi.append(expElementString)

    # joins list
    for p in range(len(eduStringLi)):
        expTempString = expTempString + "\n" + expStringLi[p]
    
    ws['F' + row].value = expTempString

    for y in range(len(education)):
        for z in range(len(education[y])):
            tinyURLString = shortener.tinyurl.short(str(education[y][z][3]))
            eduElementString = " - Institution: ", str(education[y][z][0]) + "\n    - Degree: ", str(education[y][z][1]) + "\n    - Duration: ", str(education[y][z][2]) + "\n    - URL: ", tinyURLString

            eduStringLi.append(eduElementString)

    for p in range(len(eduStringLi)):
        eduTempString = eduTempString + expStringLi[p] + "\n"
    
    print(eduTempString)
    ws['G' + row].value = eduTempString

wb.save("test2.xlsx")
