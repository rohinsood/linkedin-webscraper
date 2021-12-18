from openpyxl import *
from openpyxl.styles.colors import *
# from ExampleResults import *
from openpyxl.styles.alignment import *
from openpyxl.styles import *
from openpyxl.utils import *
from pyshorteners import *
from Scraper.Constants import *

shortener = Shortener()
expStringLi = []
eduStringLi = []
expTempString = ""
eduTempString = ""

def exporting():
    wb = load_workbook("Excel/Template.xlsx")

    ws = wb.active

    # iterate over amount of profiles
    for i in range(profileCount):
        # imports names, title, currentCOmpany, locations, and the links 1 person at a time
        row = str(i+3)
        expTempString = ""
        eduTempString = ""
        expStringLi.clear()
        eduStringLi.clear()
        
        ws['B' + row].value = names[i]
        ws['C' + row].value = currentPos[i]
        ws['D' + row].value = currentCo[i]
        ws['E' + row].value = locations[i]
        tempLinkStr = links[i]
        ws['H' + row].value = shortener.tinyurl.short(tempLinkStr)
        
        # creates a bulleted list for exerpience and edu using for loops to iterate over the different amount of experience members and experience elements that a profile may have
        for z in range(len(experience[i])):
            tinyURLString = shortener.tinyurl.short(str(experience[i][z][3]))
            expElementString = (" - Position: " + str(experience[i][z][0]) + "\n    - Company: "
            + str(experience[i][z][1]) + "\n    - Duration: " + str(experience[i][z][2]) + 
            "\n    - URL: " + tinyURLString)
            expStringLi.append(expElementString)

        print(expStringLi)
        # joins list
        for p in range(len(expStringLi)):
            expTempString = expTempString + "\n" + expStringLi[p]
        
        print(expTempString)
        ws['F' + row].value = expTempString
        expTempString = ""

        for z in range(len(education[i])):
            tinyURLString = shortener.tinyurl.short(str(education[i][z][3]))
            eduElementString = (" - Institution: " + str(education[i][z][0]) + "\n    - Degree: "
                + str(education[i][z][1]) + "\n    - Duration: " + str(education[i][z][2]) + 
                "\n    - URL: " + tinyURLString)
            eduStringLi.append(eduElementString)

        for p in range(len(eduStringLi)):
            eduTempString = eduTempString + eduStringLi[p] + "\n"
        
        # print(eduTempString)
        ws['G' + row].value = eduTempString

    wb.save("Excel/test.xlsx")
